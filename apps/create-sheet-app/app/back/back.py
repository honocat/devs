# back.py
import os
import json
from slack_sdk import WebClient
from openpyxl import load_workbook
import requests as req
import pandas as pd
from unicodedata import east_asian_width

slack_token = os.environ["SLACK_BOT_TOKEN"]
client = WebClient(token=slack_token)


def parse_message(message, channel_id):
    lines = message.strip().split("\n")

    isbn_list = []
    quantity_list = []
    detail_list = []

    for line in lines[1:-1]:
        parts = line.split()
        if len(parts) == 2:
            isbn, quantity = parts
            detail_list.append("default")
        elif len(parts) == 3:
            isbn, quantity, detail = parts
            detail_list.append(detail)
        else:
            print("Error: Invalid input format")
            client.chat_postMessage(channel=channel_id, text="入力が正しくありません。")
            return None, None, None, None

        isbn_list.append(isbn)
        quantity_list.append(quantity)

    name = lines[-1].strip()
    if name == "" or name.split()[0].isdecimal():
        client.chat_postMessage(
            channel=channel_id, text="担当者名を正しく入力してください。"
        )
        return None, None, None, None

    return isbn_list, quantity_list, detail_list, name


def create_excel_order(isbn_list, quantity_list, detail_list, name, channel_id):
    df = pd.DataFrame(
        {
            "ISBN": isbn_list,
            "Quantity": quantity_list,
            "Detail": detail_list,
            "Title": None,
            "Publisher": None,
        }
    )

    for isbn in df["ISBN"]:
        if len(isbn) == 10:
            newISBN = convertISBN(isbn)
        else:
            newISBN = isbn.replace("-", "")

        try:
            res = req.get(f"https://api.openbd.jp/v1/get?isbn={newISBN}")
        except Exception as e:
            print(f"Unexpected error: {e}")
            return

        data = res.json()
        if not data[0]:
            client.chat_postMessage(
                channel=channel_id,
                text=f"ISBN: {isbn} の書籍情報が見つかりませんでした。",
            )
            continue

        sum = 0
        title = ""
        for c in data[0]["summary"]["title"]:
            w = east_asian_width(c)
            if w == "NA" or w == "H":
                sum += 1
            else:
                sum += 2

            title += c
            if sum >= 60:
                break

        df.loc[df["ISBN"] == isbn, "Title"] = title
        if not data[0]["summary"]["publisher"]:
            df.loc[df["ISBN"] == isbn, "Publisher"] = "出版社の情報がありません"
        else:
            df.loc[df["ISBN"] == isbn, "Publisher"] = data[0]["summary"]["publisher"]
        df.loc[df["ISBN"] == isbn, "ISBN"] = newISBN

    if df["Title"].isnull().all():
        return None, None

    df.sort_values("Publisher")
    publishers = df["Publisher"].unique()

    cellPublisherName = "B5"
    cellISBN = ["F15", "F17", "F19", "F21", "F23"]
    cellTitle = ["D14", "D16", "D18", "D20", "D22"]
    cellBooks = ["C14", "C16", "C18", "C20", "C22"]
    cellDetail = "C24"
    cellName = "H27"

    outputFiles = []
    outputPublisher = []

    for publisher in publishers:
        dfInF = df[df["Publisher"] == publisher]
        n = dfInF.shape[0]
        if n % 5 != 0:
            r = int(n / 5) + 1
        else:
            r = int(n / 5)

        while n > 0:
            k = 0
            for i in range(r):
                wb = load_workbook("template.xlsx")
                ws = wb["一般注文"]
                ws[cellPublisherName] = publisher
                if dfInF.iloc[k]["Detail"] != "default":
                    ws[cellDetail] = dfInF.iloc[k]["Detail"]
                ws[cellName] = name
                for j in range(min(n, 5)):
                    ws[cellISBN[j]] = str(dfInF.iloc[k]["ISBN"])
                    ws[cellTitle[j]] = dfInF.iloc[k]["Title"]
                    ws[cellBooks[j]] = dfInF.iloc[k]["Quantity"]
                    k += 1
                    n -= 1

                if i == 0:
                    outputFileName = f"/tmp/{publisher}.xlsx"
                else:
                    outputFileName = f"/tmp/{publisher}{i}.xlsx"

                wb.save(outputFileName)
                outputFiles.append(outputFileName)
                outputPublisher.append(publisher)

    return outputFiles, outputPublisher


def convertISBN(isbn):
    isbn = "978" + isbn[0:9]
    odd = 0
    even = 0
    for i in range(12):
        if (i + 1) % 2 == 0:
            even += int(isbn[i])
        else:
            odd += int(isbn[i])
    cd = 10 - (odd + even * 3) % 10
    if cd == 10:
        cd = 0
    return isbn + str(cd)


def lambda_handler(event, context):
    print(f"Received event: {json.dumps(event)}")
    try:
        channel_id = event.get("channel")
        thread_ts = event.get("thread_ts")
        raw_text = event.get("text")

        isbn_list, quantity_list, detail_list, name = parse_message(
            raw_text, channel_id
        )
        if isbn_list is None:
            print("Failed to parse message text")
            return

        excel_files, publishers = create_excel_order(
            isbn_list, quantity_list, detail_list, name, channel_id
        )
        if excel_files is None:
            print("No data")
            return

        i = 0
        FaxNumber = ""
        for excel_file in excel_files:
            if publishers[i] != "出版社の情報がありません":
                if i != 0:
                    FaxNumber = FaxNumber + "\n"

                TempMessage = publishers[i] + ": "
                if publishers[i] in fax:
                    TempMessage = TempMessage + fax[publishers[i]]
                    FaxNumber = FaxNumber + TempMessage
                else:
                    TempMessage = TempMessage + "FAX番号が登録されていません。"
                    FaxNumber = FaxNumber + TempMessage

            with open(excel_file, "rb") as file:
                client.files_upload_v2(
                    channel=channel_id,
                    file=file,
                    filename=os.path.basename(excel_file),
                    thread_ts=thread_ts,
                )
            i += 1

        Message = f"{name}さん、発注書の作成が完了しました。\n" + FaxNumber
        client.chat_postMessage(
            channel=channel_id,
            text=Message,
        )
    except KeyError as e:
        print(f"KeyError: Missing key in message: {e}")
    except json.JSONDecodeError as e:
        print(f"JSONDecodeError: Invalid JSON in front lambda message: {e}")
    except Exception as e:
        client.chat_postMessage(
            channel=event.get("channel"),
            text="予想外のエラーが発生しました。開発者へ問い合わせてください。",
        )
        print(f"Unexpected error: {e}")

    return {"statusCode": 200, "body": json.dumps("Messages processed")}


fax = {
    # あ行
    "アイテック": "0357536121",
    "あさ出版": "0120391656",
    "アスク出版": "発注しない",
    "明日香出版社": "0120003802",
    "朝日新聞出版": "0355407845",
    "株式会社アルク": "0492572280, 買切りなので発注には注意",
    "医学書院": "0338157804",
    "医歯薬出版": "0353958563",
    "インプレス": "0484498041",
    "インプレス (発売)": "0484498041",
    "ＮＨＫ出版": "0448119436",
    "オーム社": "0332333440",
    "旺文社": "0332666048",
    # "オライリージャパン"
    # か行
    "かんき出版": "0332344421",
    "学陽書房": "0352113300",
    "学事出版": "0335189018",
    "河出書房新社": "0334040338",
    "技術評論社": "0484513847",
    "きずな出版": "0332600201",
    "協同出版": "0332330970",
    "くろしお出版": "0362612879",
    "建築資料研究社": "0339873256",
    "幻冬舎": "0354116233",
    "講談社": "0339433319",
    "弘文社": "0667024732",
    "国際コミュニケーションズ・スクール": "0335819801",
    "国際ビジネスコミュニケーション協会": "0335819801",
    "語研": "0332916749",
    "コスモピア": "0353028399",
    # さ行
    "彩図社": "0359858224",
    "三修社": "0335271411",
    "サンクチュアリ出版": "0358342508",
    "三省堂": "0426489404",
    "時事通信社": "0355652168",
    "時事通信社 (発売)": "0355652168",
    "実務教育出版": "0333551839",
    "実教出版": "0332387755",
    "ジャパンタイムズ出版": "0334538023",
    "ジー・ビー": "0332218814",
    "秀和システム": "0362643094",
    "自由国民社": "0362330780",
    "自由国民社 (発売)": "0362330780",
    "翔泳社": "0353623817",
    "祥伝社": "0332659786",
    "新星出版社": "0338310758",
    "新泉社": "0352969621",
    "新潮社": "未登録",
    "スタンダーズ": "0363806136",
    "スリーエーネットワーク": "0352752729",
    "誠文堂新光社": "0332633045",
    "青月社": "0358338664",
    "西東社": "0358003128",
    "成美堂出版": "0352068159",
    "全日本ろうあ連盟": "未登録",
    "総合法令出版": "0356235351",
    "ソシム": "0120633433",
    "ソーテック社": "0332625326",
    # た行
    "ダイヤモンド社": "0357786621",
    "高橋書店": "0359577107",
    "宝島社": "0332304794",
    "中央経済グループパブリッシング": "0332914437",
    "中央法規出版": "0338378035",
    "汐文社": "0368625202",
    "つちや書店": "0338162072",
    "電気書院": "0352599162",
    "ティーエーネットワーク": "0352810180",
    "東京リーガルマインド": "0489997591",
    "東京堂出版": "0332333746",
    "東洋館出版社": "0338239208",
    "東洋経済新報社": "0332704127",
    # な行
    "永岡書店": "未登録",
    "ナツメ社": "0332915761",
    "南江堂": "0338117230",
    "日経BP出版センター": "0484213271",
    "日経BPマーケティング": "0484213271",
    "日経BPマーケティング (発売)": "0484213271",
    "日本実業出版社": "0332680831",
    "日本評論社": "0339878590",
    "ニュートンプレス": "0339416515",
    "ネットスクール出版": "0120817708",
    # は行
    "白水社": "買切りなので発注には注意",
    "美術出版社": "0332355169",
    "富士通ラーニングメディア": "05037372080",
    "扶桑社": "未登録",
    "フローラル出版": "0362655851",
    "文藝春秋": "未登録",
    "ぺりかん社": "0338143264",
    "ベレ出版": "0120154795",
    # ま行
    "マイナビ出版": "0480482600",
    "マイナビ出版 (発売)": "0480482600",
    "マガジンハウス": "0120468127",
    "三輪書店": "0338167756",
    "明治図書出版": "05031562790",
    "メジカルビュー社": "0352282059",
    "メディカ出版": "0663985081",
    # や行
    "有斐閣": "0484658300",
    # ら行
    "ラトルズ": "0359010221",
    "リックテレコム": "0338328337",
    "リットーミュージック": "0484242299",
    # わ行
    "早稲田教育出版": "0332096248",
    "早稲田経営出版": "0352769027",
    # アルファベット
    "ＣＣＣメディアハウス": "未登録",
    "FOM出版": "05037372080",
    "Ｇａｋｋｅｎ": "0570055233",
    "Gakken": "0570055233",
    "IBCパブリッシング": "0335134512",
    "Ｊリサーチ出版": "0120061322",
    "SBクリエイティブ": "0355491211",
    "ＳＢクリエイティブ": "0355491211",
    "ＴＡＣ出版": "0352769674",
    "TAC株式会社出版事業部": "0352769674",
}
